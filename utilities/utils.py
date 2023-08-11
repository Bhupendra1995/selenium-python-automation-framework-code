import inspect
import logging
from datetime import timedelta
from datetime import datetime as dt
import softest
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):
    def find_Date(self):
        today = dt.today() + timedelta(days=5)
        day = today.day
        month = today.month
        year = today.year
        year = str(year)
        if (day < 10):
            day = '0' + str(day)
        else:
            day = str(day)
        if (month < 10):
            month = '0' + str(month)
        else:
            month = str(month)
        depart_date = day + '/' + month + '/' + year
        return depart_date
    def assertListItemText(self,list,value):
        for stop in list:
            print(stop.text)
        #     self.soft_assert(self.assertEquals, stop.text, value)
        #     if stop.text == value:
        #         print("test passed")
        #     else:
        #         print("test failed")
        # self.assert_all()

    def custom_logger(log_level=logging.DEBUG):
        # Set class/Method Name from where its called
        logger_name = inspect.stack()[1][3]

        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)

        # Create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log")

        # create formatter - how wants you wants your logs to be formatted
        formatter1 = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s : %(message)s", datefmt='%d-%b-%y %H:%M:%S')

        # add formatter to the console or file handler
        fh.setFormatter(formatter1)

        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(filename, sheet_name):
        wb = load_workbook(filename=filename)
        sh = wb[sheet_name]
        datasheet = []
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(i, j).value)
            datasheet.append(row)

        return datasheet

    def read_data_from_csv(filename):
        # create empty datalist
        datalist = []
        # open CSV file
        csvdata = open(filename, "r")
        # create CSV reader
        reader = csv.reader(csvdata)
        # skip header from csv file
        next(reader)
        # Add csv data to list
        for rows in reader:
            datalist.append(rows)

        return datalist




